# Author: GouHaoliang
# Date: 2025/6/23
# Time: 9:48

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def shift_change_stats_writer(data, header, output_path="换班统计表.xlsx", reporter=""):
    # 1、修改列名
    data.rename(columns={
        'pos_no': '编号',
        'emp_name': '姓名',
        'bucket_num': '统计桶数',
        'avg_rate': '平均扣杂率',
        'unit_price': '每桶单价',
        'amount': '金额',
        'note': '备注'
    }, inplace=True)

    # 2. 使用openpyxl处理特殊表头和汇总行
    wb = Workbook()
    ws = wb.active
    max_col = len(data.columns)  # 获取最大列数

    # 添加第一行大表头（跨所有列）
    ws.append([header])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws['A1'] = header
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 4. 写入列名和数据
    # 添加列名
    ws.append(data.columns.tolist())

    # 添加数据行
    for row in dataframe_to_rows(data, index=False, header=False):
        ws.append(row)

    # 添加汇总行
    ws.append([])  # 添加空行分隔
    total_row_idx = ws.max_row + 1

    # 合并前两格写"合计"
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=2)
    ws.cell(total_row_idx, 1, "合计")
    ws.cell(total_row_idx, 1).font = Font(bold=True)
    ws.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 计算各列的值（根据列名决定计算方式）
    for col_idx in range(3, max_col + 1):
        col_name = ws.cell(2, col_idx).value  # 获取列名（第二行）
        col_letter = ws.cell(2, col_idx).column_letter

        if col_name in ['统计桶数', '金额']:
            formula = f"=SUM({col_letter}3:{col_letter}{ws.max_row - 1})"  # 自动计算范围
        elif col_name in ['平均扣杂率']:
            formula = f"=ROUND(AVERAGE({col_letter}3:{col_letter}{ws.max_row - 1}),3)"  # 自动计算范围
        else:
            formula = ""
        ws.cell(total_row_idx, col_idx, formula)

    # 添加统计人行
    ws.append([])  # 添加空行分隔
    total_row_idx = ws.max_row + 1
    # 合并同行
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=max_col)
    ws.cell(total_row_idx, 1, f"统计人：{reporter}")
    ws.cell(total_row_idx, 1).font = Font(bold=True)
    ws.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 4. 设置全局样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用边框到所有数据单元格
    for row in ws.iter_rows(min_row=2, max_row=total_row_idx, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5. 保存最终文件
    wb.save(output_path)
    print(f"文件已生成: {output_path}")


def his_stats_writer(data, header, output_path="历史统计表.xlsx", reporter=""):

    # 1、修改列名
    data.rename(columns={
        'shift': '班次',
        'time_range': '统计时间',
        'bucket_num': '统计总桶数',
        'avg_rate': '平均扣杂率',
        'avg_price': '平均桶单价',
        'amount': '统计总金额',
        'note': '备注'
    }, inplace=True)

    # 2. 使用openpyxl处理特殊表头和汇总行
    wb = Workbook()
    ws = wb.active
    max_col = len(data.columns)  # 获取最大列数

    # 添加第一行大表头（跨所有列）
    ws.append([header])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws['A1'] = header
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 4. 写入列名和数据
    # 添加列名
    ws.append(data.columns.tolist())

    # 添加数据行
    for row in dataframe_to_rows(data, index=False, header=False):
        ws.append(row)

    # 添加汇总行
    ws.append([])  # 添加空行分隔
    total_row_idx = ws.max_row + 1

    # 合并前两格写"平均扣杂率"
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=2)
    ws.cell(total_row_idx, 1, "合计")
    ws.cell(total_row_idx, 1).font = Font(bold=True)
    ws.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 计算各列的值（根据列名决定计算方式）
    for col_idx in range(3, max_col + 1):
        col_name = ws.cell(2, col_idx).value  # 获取列名（第二行）
        col_letter = ws.cell(2, col_idx).column_letter

        if col_name in ['统计总桶数', '统计总金额']:
            formula = f"=SUM({col_letter}3:{col_letter}{ws.max_row - 1})"  # 自动计算范围
        elif col_name in ['平均扣杂率', '平均桶单价']:
            formula = f"=ROUND(AVERAGE({col_letter}3:{col_letter}{ws.max_row - 1}),3)"  # 自动计算范围
        else:
            formula = ""
        ws.cell(total_row_idx, col_idx, formula)

    # 4. 设置全局样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用边框到所有数据单元格
    for row in ws.iter_rows(min_row=2, max_row=total_row_idx, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5. 保存最终文件
    wb.save(output_path)
    print(f"文件已生成: {output_path}")


def his_detail_writer(data, header, output_path="工位历史详情表.xlsx", reporter=""):
    # 1、修改列名
    data.rename(columns={
        'op_time': '时间',
        'impurity_rate': '扣杂率',
        'note': '备注'
    }, inplace=True)
    data['序号'] = range(1, len(data) + 1)
    data = data[['序号', '时间', '扣杂率', '备注']]

    # 2. 使用openpyxl处理特殊表头和汇总行
    wb = Workbook()
    ws = wb.active
    max_col = len(data.columns)  # 获取最大列数

    # 添加第一行大表头（跨所有列）
    ws.append([header])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws['A1'] = header
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 4. 写入列名和数据
    # 添加列名
    ws.append(data.columns.tolist())

    # 添加数据行
    for row in dataframe_to_rows(data, index=False, header=False):
        ws.append(row)

    # 添加汇总行
    ws.append([])  # 添加空行分隔
    total_row_idx = ws.max_row + 1

    # 合并前两格写"平均扣杂率"
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=2)
    ws.cell(total_row_idx, 1, "平均扣杂率")
    ws.cell(total_row_idx, 1).font = Font(bold=True)
    ws.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 计算各列的值（根据列名决定计算方式）
    for col_idx in range(3, max_col + 1):
        col_name = ws.cell(2, col_idx).value  # 获取列名（第二行）
        col_letter = ws.cell(2, col_idx).column_letter

        if col_name in ['扣杂率']:
            formula = f"=ROUND(AVERAGE({col_letter}3:{col_letter}{ws.max_row - 1}),3)"  # 自动计算范围
        else:
            formula = ""
        ws.cell(total_row_idx, col_idx, formula)

    # 添加统计人行
    ws.append([])  # 添加空行分隔
    total_row_idx = ws.max_row + 1
    # 合并同行
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=max_col)
    ws.cell(total_row_idx, 1, f"统计人：{reporter}")
    ws.cell(total_row_idx, 1).font = Font(bold=True)
    ws.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 4. 设置全局样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用边框到所有数据单元格
    for row in ws.iter_rows(min_row=2, max_row=total_row_idx, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5. 保存最终文件
    wb.save(output_path)
    print(f"文件已生成: {output_path}")


if __name__ == '__main__':
    pass
